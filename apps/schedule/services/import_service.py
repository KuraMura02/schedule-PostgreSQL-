import logging
import re
from dataclasses import asdict, dataclass
from datetime import datetime, time
from io import BytesIO

from django.db import models, transaction
from openpyxl import load_workbook
from pypdf import PdfReader
from rest_framework.exceptions import ValidationError

from apps.courses.models import Course
from apps.users.models import User

from ..serializers import ScheduleSerializer


logger = logging.getLogger(__name__)


EXPECTED_HEADERS = {
    "day": {"day", "weekday"},
    "start_time": {"start_time", "start", "start time"},
    "end_time": {"end_time", "end", "end time"},
    "room": {"room", "auditorium", "classroom"},
    "teacher": {"teacher", "teacher_username", "teacher_email"},
    "course": {"course", "course_code", "course_name"},
    "student_group": {"student_group", "group", "student group"},
}


@dataclass
class ImportRowResult:
    row_number: int
    status: str
    payload: dict
    errors: dict


class BaseScheduleParser:
    def parse(self, uploaded_file):
        raise NotImplementedError


class ExcelScheduleParser(BaseScheduleParser):
    def parse(self, uploaded_file):
        workbook = load_workbook(filename=BytesIO(uploaded_file.read()), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError({"file": "The uploaded Excel file is empty."})

        header_map = self._map_headers(rows[0])
        records = []
        for row_number, row in enumerate(rows[1:], start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            record = {}
            for field, idx in header_map.items():
                record[field] = row[idx] if idx < len(row) else None
            record["_row_number"] = row_number
            records.append(record)
        return records

    def _map_headers(self, header_row):
        normalized = {self._normalize(value): idx for idx, value in enumerate(header_row) if value is not None}
        mapping = {}
        for field, variants in EXPECTED_HEADERS.items():
            for variant in variants:
                if variant in normalized:
                    mapping[field] = normalized[variant]
                    break
            else:
                if field != "student_group":
                    raise ValidationError({"file": f"Missing required column: {field}."})
        return mapping

    @staticmethod
    def _normalize(value):
        return str(value).strip().lower().replace("-", "_")


class PDFScheduleParser(BaseScheduleParser):
    separator_pattern = re.compile(r"\s*\|\s*")

    def parse(self, uploaded_file):
        reader = PdfReader(BytesIO(uploaded_file.read()))
        lines = []
        for page in reader.pages:
            text = page.extract_text() or ""
            lines.extend([line.strip() for line in text.splitlines() if line.strip()])

        records = []
        for row_number, line in enumerate(lines, start=1):
            parts = self.separator_pattern.split(line)
            if len(parts) < 6:
                continue
            records.append(
                {
                    "day": parts[0],
                    "start_time": parts[1],
                    "end_time": parts[2],
                    "room": parts[3],
                    "teacher": parts[4],
                    "course": parts[5],
                    "student_group": parts[6] if len(parts) > 6 else "",
                    "_row_number": row_number,
                }
            )

        if not records:
            raise ValidationError(
                {
                    "file": "Unable to parse PDF. Expected lines like 'mon | 09:00 | 10:30 | A-101 | teacher | CS101 | Group A'."
                }
            )
        return records


class ScheduleImportService:
    def __init__(self, uploaded_file, actor):
        self.uploaded_file = uploaded_file
        self.actor = actor
        self.parser = self._get_parser(uploaded_file.name)

    def _get_parser(self, filename):
        extension = filename.rsplit(".", 1)[-1].lower()
        if extension == "xlsx":
            return ExcelScheduleParser()
        if extension == "pdf":
            return PDFScheduleParser()
        raise ValidationError({"file": "Unsupported file type."})

    def process(self):
        self.uploaded_file.seek(0)
        raw_records = self.parser.parse(self.uploaded_file)
        results = []
        imported_ids = []

        for raw_record in raw_records:
            row_number = raw_record.pop("_row_number", 0)
            try:
                prepared = self._normalize_record(raw_record)
                # Each row is isolated to keep partial imports valid and auditable.
                with transaction.atomic():
                    serializer = ScheduleSerializer(data=prepared)
                    serializer.is_valid(raise_exception=True)
                    schedule = serializer.save()
                imported_ids.append(schedule.id)
                row_result = ImportRowResult(row_number=row_number, status="imported", payload=prepared, errors={})
                logger.info("Imported schedule row %s: %s", row_number, prepared)
            except Exception as exc:
                error_payload = getattr(exc, "detail", None) or getattr(exc, "message_dict", None) or {"detail": str(exc)}
                row_result = ImportRowResult(row_number=row_number, status="failed", payload=raw_record, errors=error_payload)
                logger.error("Failed to import schedule row %s: %s", row_number, error_payload)
            results.append(asdict(row_result))

        return {
            "summary": {
                "total_rows": len(results),
                "imported_rows": len(imported_ids),
                "failed_rows": len(results) - len(imported_ids),
                "imported_ids": imported_ids,
            },
            "rows": results,
        }

    def _normalize_record(self, record):
        teacher = self._resolve_teacher(record.get("teacher"))
        course = self._resolve_course(record.get("course"))
        return {
            "day": str(record.get("day", "")).strip().lower()[:3],
            "start_time": self._parse_time(record.get("start_time")),
            "end_time": self._parse_time(record.get("end_time")),
            "room": str(record.get("room", "")).strip(),
            "teacher": teacher.id,
            "course": course.id,
            "student_group": str(record.get("student_group", "") or "").strip(),
        }

    @staticmethod
    def _parse_time(value):
        if isinstance(value, time):
            return value
        if isinstance(value, datetime):
            return value.time()
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(str(value).strip(), fmt).time()
            except ValueError:
                continue
        raise ValidationError({"time": f"Invalid time value: {value}."})

    @staticmethod
    def _resolve_teacher(value):
        lookup_value = str(value).strip()
        teacher = (
            User.objects.filter(role=User.Role.TEACHER)
            .filter(
                models.Q(username__iexact=lookup_value)
                | models.Q(email__iexact=lookup_value)
                | models.Q(first_name__iexact=lookup_value)
            )
            .first()
        )
        if not teacher:
            raise ValidationError({"teacher": f"Teacher '{lookup_value}' was not found."})
        return teacher

    @staticmethod
    def _resolve_course(value):
        lookup_value = str(value).strip()
        course = Course.objects.filter(code__iexact=lookup_value).first() or Course.objects.filter(name__iexact=lookup_value).first()
        if not course:
            raise ValidationError({"course": f"Course '{lookup_value}' was not found."})
        return course
